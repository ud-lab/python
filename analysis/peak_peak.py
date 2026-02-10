%matplotlib widget
import pandas as pd
import numpy as np
import japanize_matplotlib
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter, find_peaks
import os
from openpyxl import Workbook

# 生データが格納されているフォルダのパスを指定してください
input_folder = '生データ/QiT_T1'
output_excel = 'CHARIoT解析_QiT_T1_5mmPMax⊿Re(TD slope).xlsx'

# 平滑化のパラメータ
window_length = 8  # 平滑化のウィンドウサイズ（奇数）
polyorder = 2      # 平滑化の多項式の次数

# ピークと谷を検出するためのパラメータ
peak_distance = 3  # ピーク間の最小距離（データポイント数）5スタンダード
peak_prominence = 0.05  # ピークの目立ち具合（プロミネンス）0.08

# TD方向の範囲
td_min = 120
td_max = 250

# ピーク幅の最大閾値（この幅以下のピークを対象にする）
max_width_threshold = 5  # 8mmスタンダード

# Excelファイルを作成
wb = Workbook()
wb.remove(wb.active)  # デフォルトのシートを削除

# 平均値をまとめるためのリスト
summary_data = []

# フォルダ内のすべてのCSVファイルを処理
for file_name in os.listdir(input_folder):
    if file_name.endswith('.csv'):
        file_path = os.path.join(input_folder, file_name)
        
        # CSVデータを読み込む
        data = pd.read_csv(file_path, header=None)
        
        # データの行数を取得
        num_rows = data.shape[0]
        
        # 真ん中の行番号を計算（偶数の場合は繰り下げ）
        middle_row_index = num_rows // 5 * 2 - 1
        #middle_row_index = num_rows // 4 - 1
        
        # 真ん中の行データを取得
        middle_row_data = data.iloc[middle_row_index, :].values
        
        # 横軸の値を0.5mm刻みで作成
        x_values = np.arange(0, len(middle_row_data) * 0.5, 0.5)
        
        # TD方向の範囲でデータを切り取る
        mask = (x_values >= td_min) & (x_values <= td_max)
        x_values = x_values[mask]
        middle_row_data = middle_row_data[mask]
        
        # 平滑化（Savitzky-Golayフィルタを使用）
        smoothed_data = savgol_filter(middle_row_data, window_length=window_length, polyorder=polyorder)
        
        # ピークと谷を検出
        peaks, peak_properties = find_peaks(middle_row_data, distance=peak_distance, prominence=peak_prominence)
        troughs, trough_properties = find_peaks(-middle_row_data, distance=peak_distance, prominence=peak_prominence)
        
        # ピークと谷の情報を取得
        peak_positions = x_values[peaks]
        peak_intensities = middle_row_data[peaks]
        trough_positions = x_values[troughs]
        trough_intensities = middle_row_data[troughs]
        
        # ピークと谷の対応付け
        wave_heights = []
        wave_pitches = []
        peak_widths = []
        peak_slopes = []
        filtered_peak_heights = []  # 幅が閾値以下のピークの高さを格納するリスト
        filtered_peak_slopes = []
        for i in range(len(peaks)):
            peak_idx = peaks[i]
            
            # 前後の谷を探す
            prev_trough_idx = troughs[troughs < peak_idx]
            next_trough_idx = troughs[troughs > peak_idx]
            
            # 前後の谷が存在する場合、波の高さを計算
            if len(prev_trough_idx) > 0 and len(next_trough_idx) > 0:
                prev_trough = middle_row_data[prev_trough_idx[-1]]
                next_trough = middle_row_data[next_trough_idx[0]]
                wave_height = max(middle_row_data[peak_idx] - prev_trough, middle_row_data[peak_idx] - next_trough)
                wave_heights.append(wave_height)
                wave_pitch = x_values[next_trough_idx[0]] - x_values[prev_trough_idx[-1]]
                wave_pitches.append(wave_pitch)
                peak_width = x_values[next_trough_idx[0]] - x_values[prev_trough_idx[-1]]
                peak_widths.append(peak_width)
                peak_slope = np.abs(wave_height / peak_width)
                peak_slopes.append(peak_slope)
                
                # 幅が閾値以下の場合、ピークの高さを記録
                if peak_width <= max_width_threshold:
                    filtered_peak_heights.append(wave_height)
                    filtered_peak_slopes.append(peak_slope)
        
        # 幅が閾値以下のピークの高さの平均を計算
        #average_filtered_peak_height = np.mean(filtered_peak_heights) if filtered_peak_heights else 0
        max_filtered_peak_slope = np.amax(filtered_peak_slopes) if filtered_peak_slopes else 0
        # 最大値をまとめるリストに追加
        summary_data.append([file_name, max_filtered_peak_slope])
        
        # 結果をExcelに書き込む
        sheet_title = file_name[:10]  # タイトルをファイル名の最初10文字に設定
        sheet = wb.create_sheet(title=sheet_title)
        sheet.append(['TD Position [mm]', 'Peak Intensity', 'Position [mm]', 'Trough Intensity', 'Pitch (mm)', 'Height', 'Width', 'Slope'])
        for i in range(len(peaks)):
            peak_pos = peak_positions[i] if i < len(peak_positions) else None
            peak_int = peak_intensities[i] if i < len(peak_intensities) else None
            trough_pos = trough_positions[i] if i < len(trough_positions) else None
            trough_int = trough_intensities[i] if i < len(trough_intensities) else None
            pitch = wave_pitches[i] if i < len(wave_pitches) else None
            height = wave_heights[i] if i < len(wave_heights) else None
            width = peak_widths[i] if i < len(peak_widths) else None
            slope = peak_slopes[i] if i < len(peak_slopes) else None
            sheet.append([peak_pos, peak_int, trough_pos, trough_int, pitch, height, width, slope])
        
        sheet.append([])
        sheet.append(['Max Slope of Peaks with Width <= Threshold', max_filtered_peak_slope])
        
        # ピーク幅の分布をプロット
        plt.figure(figsize=(10, 6))
        plt.hist(peak_widths, bins=20, alpha=0.7, color='blue', edgecolor='black')
        plt.title(f'{sheet_title} - Peak Width Distribution')
        plt.xlabel('Peak Width [mm]')
        plt.ylabel('Frequency')
        plt.grid(True)
        plt.show()
        
        # グラフをプロット
        plt.figure(figsize=(10, 6))
        plt.plot(x_values, smoothed_data, label='Smoothed Data', linewidth=1, alpha=0.5)
        plt.plot(x_values, middle_row_data, label='Original Data')
        plt.plot(peak_positions, peak_intensities, 'ro', label='ムラの山')
        plt.plot(trough_positions, trough_intensities, 'bo', label='ムラの谷')
        plt.title(f'{sheet_title}')
        plt.xlabel('TD Position [mm]')
        plt.ylabel('Re [nm]')
        
        plt.grid(True)
        plt.legend()
        plt.show()

# 新しいシートにまとめを記録
summary_sheet = wb.create_sheet(title="Summary")
summary_sheet.append(['File Name', 'Max Slope of Peaks with Width <= Threshold'])
for row in summary_data:
    summary_sheet.append(row)

# Excelファイルを保存
wb.save(output_excel)
print(f"解析結果を {output_excel} に保存しました。")
